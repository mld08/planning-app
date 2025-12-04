import base64
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail
from flask_migrate import Migrate
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta, date
import os
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Config
from models import db, User, Planning, Affectation, HistoriqueModification
from scheduler import PlanningScheduler

# Initialiser l'application Flask
app = Flask(__name__)
app.config.from_object(Config)
application = app  

# Initialiser les extensions
db.init_app(app)
migrate = Migrate(app, db)
mail = Mail(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter pour accéder à cette page.'

# Initialiser le scheduler
scheduler_manager = PlanningScheduler(app, mail)
scheduler = BackgroundScheduler()


@login_manager.user_loader
def load_user(user_id):
    """Charge un utilisateur depuis la base de données"""
    return User.query.get(int(user_id))



# ========================== HELPER FUNCTIONS ==========================

def _number_to_roman(num):
    """Convertir un nombre en chiffres romains"""
    val = [1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1]
    syms = ["M", "CM", "D", "CD", "C", "XC", "L", "XL", "X", "IX", "V", "IV", "I"]
    roman_num = ''
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman_num += syms[i]
            num -= val[i]
        i += 1
    return roman_num


# def _organiser_affectations_par_sections(planning):
#     """
#     Organise les affectations selon les 12 sections du template
#     Retourne un dictionnaire structuré par section
#     """
#     affectations = Affectation.query.filter_by(planning_id=planning.id).all()
    
#     # Structure pour stocker les données organisées
#     sections_data = {
#         'I_VEILLE_CRSS': {
#             'titre': 'I - VEILLE CRSS',
#             'couleur': '70AD47',  # Vert
#             'sous_sections': {
#                 'journee': {'titre': 'Journée:  08h - 17h', 'couleur': 'FFFF00', 'agents': {}},
#                 'nuit': {'titre': 'Nuit : 17h - 08h', 'couleur': '00B050', 'agents': {}}
#             }
#         },
#         'II_BRIGADE_BVP': {
#             'titre': 'II - BRIGADE DE VEILLE PORTUAIRE',
#             'couleur': 'FFC000',  # Orange
#             'postes': {
#                 'chef': {'titre': 'Chef d\'equipe', 'agents': {}},
#                 'inspecteur': {'titre': 'Inspecteur de Veille', 'agents': {}},
#                 'autres': {'titre': 'Autres agents', 'agents': {}},
#                 'chauffeurs': {'titre': 'Chauffeurs', 'couleur': 'FFFF00', 'agents': {}}
#             }
#         },
#         'III_REPOS_CHAUFFEURS': {
#             'titre': 'III - REPOS CHAUFFEURS',
#             'couleur': '92D050',  # Vert clair
#             'chauffeurs': {'titre': 'Chauffeurs', 'agents': {}}
#         },
#         'IV_CERTIFICATION_DIR': {
#             'titre': 'IV - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - DIRECTION',
#             'couleur': '00B0F0',  # Bleu clair
#             'agents': {}
#         },
#         'V_CERTIFICATION_AERO': {
#             'titre': 'V - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - AEROPORT',
#             'couleur': '00B050',  # Vert
#             'agents': {}
#         },
#         'VI_PATROUILLE_MARITIME': {
#             'titre': 'VI - PATROUILLE MARITIME COTIERE',
#             'couleur': '0070C0',  # Bleu
#             'inspecteurs': {'titre': 'Inspecteurs', 'agents': {}}
#         },
#         'VII_GARDIENNAGE': {
#             'titre': 'VII - GARDIENNAGE',
#             'couleur': '7030A0',  # Violet
#             'agents': {}
#         },
#         'VIII_COURRIER': {
#             'titre': 'VIII - COURRIER',
#             'couleur': 'FFC000',  # Orange
#             'chauffeurs': {'titre': 'Chauffeurs', 'couleur': 'FFFF00', 'agents': {}}
#         },
#         'IX_INSPECTION_USINES': {
#             'titre': 'IX - INSPECTION USINES',
#             'couleur': '00B050',  # Vert
#             'postes': {
#                 'inspecteurs': {'titre': 'Inspecteurs', 'agents': {}},
#                 'chauffeurs': {'titre': 'Chauffeurs', 'couleur': 'FFFF00', 'agents': {}}
#             }
#         },
#         'X_PATROUILLE_AERIENNE': {
#             'titre': 'X - PATROUILLE AERIENNE',
#             'couleur': '4472C4',  # Bleu
#             'inspecteurs': {'titre': 'Inspecteurs', 'agents': {}}
#         },
#         'XI_MISSION_CONGES': {
#             'titre': 'XI - EN MISSION OU EN CONGES',
#             'couleur': 'C00000',  # Rouge
#             'format_special': 'periode',
#             'personnes': []  # Liste de {nom, periode}
#         },
#         'XII_EMBARQUEMENTS': {
#             'titre': 'XII -  EMBARQUEMENTS',
#             'couleur': '002060',  # Bleu foncé
#             'format_special': 'liste',
#             'observateurs': []  # Liste numérotée
#         }
#     }
    
#     # Organiser les affectations dans les bonnes sections
#     for aff in affectations:
#         agent = aff.agent
#         jour_index = (aff.jour - planning.date_debut).days
        
#         # Section I - VEILLE CRSS
#         if aff.equipe == 'CRSS':
#             if aff.shift == 'jour':
#                 if agent.id not in sections_data['I_VEILLE_CRSS']['sous_sections']['journee']['agents']:
#                     sections_data['I_VEILLE_CRSS']['sous_sections']['journee']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['I_VEILLE_CRSS']['sous_sections']['journee']['agents'][agent.id]['jours'][jour_index] = aff
#             elif aff.shift == 'nuit':
#                 if agent.id not in sections_data['I_VEILLE_CRSS']['sous_sections']['nuit']['agents']:
#                     sections_data['I_VEILLE_CRSS']['sous_sections']['nuit']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['I_VEILLE_CRSS']['sous_sections']['nuit']['agents'][agent.id]['jours'][jour_index] = aff
        
#         # Section II - BRIGADE BVP
#         elif aff.equipe == 'BVP':
#             poste_key = aff.poste if aff.poste in ['chef', 'inspecteur'] else 'autres'
            
#             # Vérifier si c'est un chauffeur
#             if agent.fonction and 'chauffeur' in agent.fonction.lower():
#                 poste_key = 'chauffeurs'
            
#             if agent.id not in sections_data['II_BRIGADE_BVP']['postes'][poste_key]['agents']:
#                 sections_data['II_BRIGADE_BVP']['postes'][poste_key]['agents'][agent.id] = {
#                     'nom': agent.nom_complet,
#                     'jours': {}
#                 }
#             sections_data['II_BRIGADE_BVP']['postes'][poste_key]['agents'][agent.id]['jours'][jour_index] = aff
        
#         # Déterminer les autres sections selon les attributs de l'agent ou le poste
#         elif aff.poste:
#             poste_lower = aff.poste.lower()
            
#             # Section IV - Certification Direction
#             if 'certification' in poste_lower and 'direction' in poste_lower:
#                 if agent.id not in sections_data['IV_CERTIFICATION_DIR']['agents']:
#                     sections_data['IV_CERTIFICATION_DIR']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['IV_CERTIFICATION_DIR']['agents'][agent.id]['jours'][jour_index] = aff
            
#             # Section V - Certification Aéroport
#             elif agent.est_certification_aeroport or ('certification' in poste_lower and 'aero' in poste_lower):
#                 if agent.id not in sections_data['V_CERTIFICATION_AERO']['agents']:
#                     sections_data['V_CERTIFICATION_AERO']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['V_CERTIFICATION_AERO']['agents'][agent.id]['jours'][jour_index] = aff
            
#             # Section VI - Patrouille Maritime
#             elif 'patrouille' in poste_lower and 'maritime' in poste_lower:
#                 if agent.id not in sections_data['VI_PATROUILLE_MARITIME']['inspecteurs']['agents']:
#                     sections_data['VI_PATROUILLE_MARITIME']['inspecteurs']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['VI_PATROUILLE_MARITIME']['inspecteurs']['agents'][agent.id]['jours'][jour_index] = aff
            
#             # Section VII - Gardiennage
#             elif 'gardien' in poste_lower:
#                 if agent.id not in sections_data['VII_GARDIENNAGE']['agents']:
#                     sections_data['VII_GARDIENNAGE']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['VII_GARDIENNAGE']['agents'][agent.id]['jours'][jour_index] = aff
            
#             # Section VIII - Courrier (Chauffeurs)
#             elif 'courrier' in poste_lower:
#                 if agent.id not in sections_data['VIII_COURRIER']['chauffeurs']['agents']:
#                     sections_data['VIII_COURRIER']['chauffeurs']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['VIII_COURRIER']['chauffeurs']['agents'][agent.id]['jours'][jour_index] = aff
            
#             # Section IX - Inspection Usines
#             elif 'inspection' in poste_lower and 'usine' in poste_lower:
#                 if agent.fonction and 'chauffeur' in agent.fonction.lower():
#                     if agent.id not in sections_data['IX_INSPECTION_USINES']['postes']['chauffeurs']['agents']:
#                         sections_data['IX_INSPECTION_USINES']['postes']['chauffeurs']['agents'][agent.id] = {
#                             'nom': agent.nom_complet,
#                             'jours': {}
#                         }
#                     sections_data['IX_INSPECTION_USINES']['postes']['chauffeurs']['agents'][agent.id]['jours'][jour_index] = aff
#                 else:
#                     if agent.id not in sections_data['IX_INSPECTION_USINES']['postes']['inspecteurs']['agents']:
#                         sections_data['IX_INSPECTION_USINES']['postes']['inspecteurs']['agents'][agent.id] = {
#                             'nom': agent.nom_complet,
#                             'jours': {}
#                         }
#                     sections_data['IX_INSPECTION_USINES']['postes']['inspecteurs']['agents'][agent.id]['jours'][jour_index] = aff
            
#             # Section X - Patrouille Aérienne
#             elif 'patrouille' in poste_lower and 'aer' in poste_lower:
#                 if agent.id not in sections_data['X_PATROUILLE_AERIENNE']['inspecteurs']['agents']:
#                     sections_data['X_PATROUILLE_AERIENNE']['inspecteurs']['agents'][agent.id] = {
#                         'nom': agent.nom_complet,
#                         'jours': {}
#                     }
#                 sections_data['X_PATROUILLE_AERIENNE']['inspecteurs']['agents'][agent.id]['jours'][jour_index] = aff
    
#     # Section XI - Mission ou Congés (basé sur les attributs User)
#     # Chercher les agents en mission/congés pendant cette semaine
#     tous_agents = User.query.filter_by(role='agent').all()
#     for agent in tous_agents:
#         if not agent.disponibilite:  # Agent non disponible = en mission ou congés
#             sections_data['XI_MISSION_CONGES']['personnes'].append({
#                 'nom': agent.nom_complet,
#                 'periode': 'EN MISSION'  # Vous pouvez ajouter des dates si disponibles
#             })
    
#     # Section XII - Embarquements (observateurs embarqués)
#     for agent in tous_agents:
#         if agent.est_observateur_embarque:
#             # Vérifier si embarqué pendant cette semaine
#             if agent.date_embarquement and agent.date_debarquement_prevue:
#                 if (agent.date_embarquement <= planning.date_fin and 
#                     agent.date_debarquement_prevue >= planning.date_debut):
#                     sections_data['XII_EMBARQUEMENTS']['observateurs'].append({
#                         'nom': agent.nom_complet,
#                         'date_emb': agent.date_embarquement,
#                         'date_deb': agent.date_debarquement_prevue
#                     })
    
#     return sections_data


def _organiser_affectations_par_sections(planning):
    """
    Organise les affectations selon les 12 sections du template
    Retourne un dictionnaire structuré par section (compatible avec l'export Excel/PDF).
    """

    affectations = Affectation.query.filter_by(planning_id=planning.id).all()
    tous_agents = User.query.filter_by(role='agent').all()

    # Structure initiale complète (mêmes clés attendues par l'export)
    sections_data = {
        'I_VEILLE_CRSS': {
            'titre': 'I - VEILLE CRSS',
            'couleur': '70AD47',
            'sous_sections': {
                'journee': {'titre': 'Journée:  08h - 17h', 'couleur': 'FFFF00', 'agents': {}},
                'nuit': {'titre': 'Nuit : 17h - 08h', 'couleur': '00B050', 'agents': {}}
            }
        },
        'II_BRIGADE_BVP': {
            'titre': 'II - BRIGADE DE VEILLE PORTUAIRE',
            'couleur': 'FFC000',
            'postes': {
                'chef': {'titre': "Chef d'équipe", 'agents': {}},
                'inspecteur': {'titre': 'Inspecteur de Veille', 'agents': {}},
                'autres': {'titre': 'Autres agents', 'agents': {}},
                'chauffeurs': {'titre': 'Chauffeurs', 'couleur': 'FFFF00', 'agents': {}}
            }
        },
        'III_REPOS_CHAUFFEURS': {
            'titre': 'III - REPOS CHAUFFEURS',
            'couleur': '92D050',
            'chauffeurs': {'titre': 'Chauffeurs', 'agents': {}}
        },
        'IV_CERTIFICATION_DIR': {
            'titre': 'IV - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - DIRECTION',
            'couleur': '00B0F0',
            'agents': {}
        },
        'V_CERTIFICATION_AERO': {
            'titre': 'V - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - AEROPORT',
            'couleur': '00B050',
            'agents': {}
        },
        'VI_PATROUILLE_MARITIME': {
            'titre': 'VI - PATROUILLE MARITIME COTIERE',
            'couleur': '0070C0',
            'inspecteurs': {'titre': 'Inspecteurs', 'agents': {}}
        },
        'VII_GARDIENNAGE': {
            'titre': 'VII - GARDIENNAGE',
            'couleur': '7030A0',
            'agents': {}
        },
        'VIII_COURRIER': {
            'titre': 'VIII - COURRIER',
            'couleur': 'FFC000',
            'chauffeurs': {'titre': 'Chauffeurs', 'couleur': 'FFFF00', 'agents': {}}
        },
        'IX_INSPECTION_USINES': {
            'titre': 'IX - INSPECTION USINES',
            'couleur': '00B050',
            'postes': {
                'inspecteurs': {'titre': 'Inspecteurs', 'agents': {}},
                'chauffeurs': {'titre': 'Chauffeurs', 'couleur': 'FFFF00', 'agents': {}}
            }
        },
        'X_PATROUILLE_AERIENNE': {
            'titre': 'X - PATROUILLE AERIENNE',
            'couleur': '4472C4',
            'inspecteurs': {'titre': 'Inspecteurs', 'agents': {}}
        },
        'XI_MISSION_CONGES': {
            'titre': 'XI - EN MISSION OU EN CONGES',
            'couleur': 'C00000',
            'format_special': 'periode',
            'personnes': []  # Liste de {nom, periode}
        },
        'XII_EMBARQUEMENTS': {
            'titre': 'XII -  EMBARQUEMENTS',
            'couleur': '002060',
            'format_special': 'liste',
            'observateurs': []  # Liste numérotée: {nom, date_emb, date_deb}
        }
    }

    # -------------------------
    # 1) Chauffeurs en repos (disponibilite == False) -> section III
    # -------------------------
    for agent in tous_agents:
        if agent.est_chauffeur and not agent.disponibilite:
            # ajouter comme chauffeur en repos (sans jours - section "personne")
            sections_data['III_REPOS_CHAUFFEURS']['chauffeurs']['agents'][agent.id] = {
                'nom': agent.nom_complet,
                'jours': {}  # pas d'affectations, liste vide
            }

    # -------------------------
    # 2) Organiser les affectations normales
    # -------------------------
    for aff in affectations:
        agent = aff.agent
        # calcul indice du jour (0..6)
        jour_index = (aff.jour - planning.date_debut).days
        if jour_index < 0 or jour_index > 6:
            # en dehors de la semaine : ignorer
            continue

        # Si l'agent est chauffeur mais en repos -> on l'ignore (déjà listé en III)
        if agent.est_chauffeur and not agent.disponibilite:
            continue

        # Section I - VEILLE CRSS
        if aff.equipe == 'CRSS':
            key = 'journee' if aff.shift == 'jour' else 'nuit'
            target = sections_data['I_VEILLE_CRSS']['sous_sections'][key]['agents']
            if agent.id not in target:
                target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
            target[agent.id]['jours'][jour_index] = aff

        # Section II - BRIGADE BVP
        elif aff.equipe == 'BVP':
            # si l'agent est chauffeur => mettre dans chauffeurs BVP
            if agent.est_chauffeur:
                poste_container = sections_data['II_BRIGADE_BVP']['postes']['chauffeurs']['agents']
                print("Agent chauffeur BVP:", agent.nom_complet)
            else:
                # poste peut être 'chef' ou 'inspecteur' sinon 'autres'
                p = aff.poste if aff.poste in ['chef', 'inspecteur'] else 'autres'
                poste_container = sections_data['II_BRIGADE_BVP']['postes'][p]['agents']

            if agent.id not in poste_container:
                poste_container[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
            poste_container[agent.id]['jours'][jour_index] = aff

        # Section VIII - Courrier (si poste 'courrier' et chauffeur)
        elif (aff.poste or '').lower() == 'courrier' and agent.est_chauffeur:
            target = sections_data['VIII_COURRIER']['chauffeurs']['agents']
            if agent.id not in target:
                target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
            target[agent.id]['jours'][jour_index] = aff

        # Section IX - Inspection Usines (si poste contient 'usine')
        elif aff.poste and 'usine' in aff.poste.lower():
            if agent.est_chauffeur:
                target = sections_data['IX_INSPECTION_USINES']['postes']['chauffeurs']['agents']
            else:
                target = sections_data['IX_INSPECTION_USINES']['postes']['inspecteurs']['agents']
            if agent.id not in target:
                target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
            target[agent.id]['jours'][jour_index] = aff

        # Section IV / V / VI / VII / X - autres postes déterminés par poste ou attribut agent
        else:
            poste_lower = (aff.poste or '').lower()

            # IV - Certification Direction (poste contenant 'certification' et 'direction')
            if 'certification' in poste_lower and 'direction' in poste_lower:
                target = sections_data['IV_CERTIFICATION_DIR']['agents']
                if agent.id not in target:
                    target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
                target[agent.id]['jours'][jour_index] = aff

            # V - Certification Aéroport (attribut ou poste)
            elif agent.est_certification_aeroport or ('certification' in poste_lower and 'aero' in poste_lower):
                target = sections_data['V_CERTIFICATION_AERO']['agents']
                if agent.id not in target:
                    target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
                target[agent.id]['jours'][jour_index] = aff

            # VI - Patrouille maritime
            elif 'patrouille' in poste_lower and 'maritime' in poste_lower:
                target = sections_data['VI_PATROUILLE_MARITIME']['inspecteurs']['agents']
                if agent.id not in target:
                    target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
                target[agent.id]['jours'][jour_index] = aff

            # VII - Gardiennage
            elif 'gardien' in poste_lower:
                target = sections_data['VII_GARDIENNAGE']['agents']
                if agent.id not in target:
                    target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
                target[agent.id]['jours'][jour_index] = aff

            # X - Patrouille aérienne
            elif 'patrouille' in poste_lower and 'aer' in poste_lower:
                target = sections_data['X_PATROUILLE_AERIENNE']['inspecteurs']['agents']
                if agent.id not in target:
                    target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
                target[agent.id]['jours'][jour_index] = aff

            # Si rien ne matche, on peut tenter de placer selon fonction textuelle
            else:
                # si fonction mentionne 'chauffeur' et équipe non BVP, on essaye COURRIER or IX
                fn = (agent.fonction or '').lower()
                if 'chauffeur' in fn:
                    # prioriser courrier si poste contient 'courrier'
                    if 'courrier' in poste_lower:
                        target = sections_data['VIII_COURRIER']['chauffeurs']['agents']
                    else:
                        target = sections_data['II_BRIGADE_BVP']['postes']['chauffeurs']['agents']
                    if agent.id not in target:
                        target[agent.id] = {'nom': agent.nom_complet, 'jours': {}}
                    target[agent.id]['jours'][jour_index] = aff
                else:
                    # Par défaut, rien (ou on pourrait l'ajouter à 'autres' BVP si pertinent)
                    pass

    # -------------------------
    # 3) SECTION XI - Mission / Congés
    # On parcourt tous les agents et si disponibilite False on ajoute à 'personnes'
    # -------------------------
    for agent in tous_agents:
        if not agent.disponibilite:
            periode = 'EN MISSION'  # placeholder ; tu peux remplacer par les vraies dates si dispo
            sections_data['XI_MISSION_CONGES']['personnes'].append({
                'nom': agent.nom_complet,
                'periode': periode
            })

    # -------------------------
    # 4) SECTION XII - Embarquements (observateurs)
    # -------------------------
    for agent in tous_agents:
        if agent.est_observateur_embarque and agent.date_embarquement and agent.date_debarquement_prevue:
            if (agent.date_embarquement <= planning.date_fin and agent.date_debarquement_prevue >= planning.date_debut):
                sections_data['XII_EMBARQUEMENTS']['observateurs'].append({
                    'nom': agent.nom_complet,
                    'date_emb': agent.date_embarquement,
                    'date_deb': agent.date_debarquement_prevue
                })

    return sections_data


# ==================== ROUTES D'AUTHENTIFICATION ====================

@app.route('/')
def index():
    """Page d'accueil - redirige vers le dashboard approprié"""
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('dashboard_admin'))
        else:
            return redirect(url_for('dashboard_agent'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Page de connexion"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember', False)
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            next_page = request.args.get('next')
            
            flash(f'Bienvenue {user.prenom} !', 'success')
            
            if next_page:
                return redirect(next_page)
            elif user.role == 'admin':
                return redirect(url_for('dashboard_admin'))
            else:
                return redirect(url_for('dashboard_agent'))
        else:
            flash('Nom d\'utilisateur ou mot de passe incorrect.', 'danger')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Page d'inscription (réservée aux admins ou première installation)"""
    premier_utilisateur = User.query.count() == 0
    
    if not premier_utilisateur and (not current_user.is_authenticated or current_user.role != 'admin'):
        flash('Seuls les administrateurs peuvent créer des comptes.', 'danger')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        nom = request.form.get('nom')
        prenom = request.form.get('prenom')
        password = request.form.get('password')
        username = request.form.get('username')
        phone = request.form.get('phone')
        role = request.form.get('role', 'agent')
        
        if User.query.filter_by(username=username).first():
            flash('Cet username est déjà utilisé.', 'danger')
        elif username and User.query.filter_by(username=username).first():
            flash('Ce nom d\'utilisateur est déjà utilisé.', 'danger')
        else:
            if premier_utilisateur:
                role = 'admin'
            
            user = User(
                nom=nom,
                prenom=prenom,
                username=username,
                phone=phone,
                role=role
            )
            user.set_password(password)
            
            db.session.add(user)
            db.session.commit()
            
            flash(f'Compte créé avec succès pour {prenom} {nom}!', 'success')
            
            if premier_utilisateur:
                login_user(user)
                return redirect(url_for('dashboard_admin'))
            else:
                return redirect(url_for('gestion_agents'))
    
    return render_template('register.html', premier_utilisateur=premier_utilisateur)


@app.route('/logout')
@login_required
def logout():
    """Déconnexion"""
    logout_user()
    flash('Vous êtes déconnecté.', 'info')
    return redirect(url_for('login'))


# ==================== DASHBOARD ADMIN ====================

@app.route('/admin/dashboard')
@login_required
def dashboard_admin():
    """Dashboard administrateur"""
    if current_user.role != 'admin':
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard_agent'))
    
    # Statistiques
    nb_agents = User.query.filter_by(role='agent').count()
    nb_agents_disponibles = User.query.filter_by(role='agent', disponibilite=True).count()
    
    # Planning actuel
    aujourd_hui = date.today()
    planning_actuel = Planning.query.filter(
        Planning.date_debut <= aujourd_hui,
        Planning.date_fin >= aujourd_hui
    ).first()
    
    # Plannings récents
    plannings_recents = Planning.query.order_by(Planning.date_debut.desc()).limit(5).all()
    
    return render_template('dashboard_admin.html',
                         nb_agents=nb_agents,
                         nb_agents_disponibles=nb_agents_disponibles,
                         planning_actuel=planning_actuel,
                         plannings_recents=plannings_recents)


# ==================== DASHBOARD AGENT ====================

@app.route('/agent/dashboard')
@login_required
def dashboard_agent():
    """Dashboard agent"""
    aujourd_hui = date.today()
    planning_actuel = Planning.query.filter(
        Planning.date_debut <= aujourd_hui,
        Planning.date_fin >= aujourd_hui
    ).first()
    
    affectations_semaine = []
    if planning_actuel:
        affectations_semaine = Affectation.query.filter_by(
            planning_id=planning_actuel.id,
            agent_id=current_user.id
        ).order_by(Affectation.jour).all()
    
    total_jours = current_user.compteur_jour
    total_nuits = current_user.compteur_nuit
    
    return render_template('dashboard_agent.html',
                         planning_actuel=planning_actuel,
                         affectations_semaine=affectations_semaine,
                         total_jours=total_jours,
                         total_nuits=total_nuits, 
                         timedelta=timedelta)


# ==================== GESTION DES AGENTS ====================

@app.route('/admin/agents')
@login_required
def gestion_agents():
    """Liste paginée des agents"""
    if current_user.role != 'admin':
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard_agent'))
    
    # Récupère le numéro de page depuis l’URL (par défaut 1)
    page = request.args.get('page', 1, type=int)
    per_page = 100  # nombre d'agents par page (modifiable)
    
    # Pagination
    pagination = User.query.filter_by(role='agent').paginate(page=page, per_page=per_page)
    agents = pagination.items  # agents de la page courante
    all_agents = User.query.filter_by(role='agent').all()
    
    return render_template('gestion_agents.html', agents=agents, pagination=pagination, all_agents=all_agents)


@app.route('/admin/agents/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_agent():
    """Ajouter un agent"""
    if current_user.role != 'admin':
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard_agent'))
    
    if request.method == 'POST':
        nom = request.form.get('nom')
        prenom = request.form.get('prenom')
        username = request.form.get('username') or None
        phone = request.form.get('phone') or ''
        password = request.form.get('password') or 'defaultpassword'
        fonction = request.form.get('fonction')
        
        # ========== NOUVEAUX CHAMPS POUR LES CONTRAINTES ==========
        genre = request.form.get('genre')
        est_chef_bureau = request.form.get('est_chef_bureau') == 'on'
        est_certification_aeroport = request.form.get('est_certification_aeroport') == 'on'
        est_chef_equipe_bvp = request.form.get('est_chef_equipe_bvp') == 'on'
        est_chef_equipe_usine = request.form.get('est_chef_equipe_usine') == 'on'
        est_chauffeur = request.form.get('est_chauffeur') == 'on'
        est_absent = request.form.get('est_absent') == 'on'
        est_operateur_veille_crss = request.form.get('est_operateur_veille_crss') == 'on'
        date_embarquement = None
        date_debarquement_prevue = None
        aeroport = request.form.get('aeroport') == 'on'
        patrouille_cotiere = request.form.get('patrouille_cotiere') == 'on'
        inspection_usine = request.form.get('inspection_usine') == 'on'
        patrouille_aerienne = request.form.get('patrouille_aerienne') == 'on'
        gardiennage = request.form.get('gardiennage') == 'on'
        courrier = request.form.get('courrier') == 'on'
        missions = request.form.get('missions') == 'on'
        crss = request.form.get('crss') == 'on'
        bvp = request.form.get('bvp') == 'on'
        certification_dpsp = request.form.get('certification_dpsp') == 'on'
        certification = request.form.get('certification') == 'on'

        # ===========================================================

        if username and User.query.filter_by(username=username).first():
            flash('Ce nom d\'utilisateur est déjà utilisé.', 'danger')
        else:
            agent = User(
                nom=nom,
                prenom=prenom,
                username=username,
                phone=phone,
                fonction=fonction,
                role='agent',
                # Nouveaux champs contraintes
                genre=genre,
                est_chef_bureau=est_chef_bureau,
                est_certification_aeroport=est_certification_aeroport,
                est_chef_equipe_bvp=est_chef_equipe_bvp,
                est_chef_equipe_usine=est_chef_equipe_usine,
                est_chauffeur=est_chauffeur,
                est_absent=est_absent,
                est_operateur_veille_crss=est_operateur_veille_crss,
                date_embarquement=date_embarquement,
                date_debarquement_prevue=date_debarquement_prevue,
                #aeroport=aeroport,
                patrouille_cotiere=patrouille_cotiere,
                inspection_usine=inspection_usine,
                patrouille_aerienne=patrouille_aerienne,
                gardiennage=gardiennage,
                courrier=courrier,
                missions=missions,
                crss=crss,
                bvp=bvp,
                certification_dpsp=certification_dpsp,
                certification=certification
            )
            agent.set_password(password)
            
            db.session.add(agent)
            db.session.commit()
            
            flash(f'Agent {prenom} {nom} ajouté avec succès!', 'success')
            return redirect(url_for('gestion_agents'))
    
    return render_template('ajouter_agent.html')


@app.route('/admin/agents/<int:agent_id>/modifier', methods=['GET', 'POST'])
@login_required
def modifier_agent(agent_id):
    """Modifier un agent"""
    if current_user.role != 'admin':
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard_agent'))
    
    agent = User.query.get_or_404(agent_id)
    
    if request.method == 'POST':
        agent.nom = request.form.get('nom')
        agent.prenom = request.form.get('prenom')
        agent.username = request.form.get('username') or None
        agent.phone = request.form.get('phone') or ''
        agent.fonction = request.form.get('fonction') 
        agent.chef_de_mission = request.form.get('chef_de_mission')
        agent.disponibilite = request.form.get('disponibilite') == 'on'
        
        # ========== NOUVEAUX CHAMPS CONTRAINTES ==========
        agent.genre = request.form.get('genre')
        agent.est_absent = request.form.get('est_absent') == 'on'
        agent.est_chef_equipe = request.form.get('est_chef_equipe') == 'on'
        agent.est_chef_bureau = request.form.get('est_chef_bureau') == 'on'
        agent.est_certification_aeroport = request.form.get('est_certification_aeroport') == 'on'
        agent.est_chef_equipe_bvp = request.form.get('est_chef_equipe_bvp') == 'on'
        agent.est_chef_equipe_usine = request.form.get('est_chef_equipe_usine') == 'on'
        agent.est_observateur_embarque = request.form.get('est_observateur_embarque') == 'on'
        agent.est_chauffeur = request.form.get('est_chauffeur') == 'on'
        agent.est_operateur_veille_crss = request.form.get('est_operateur_veille_crss') == 'on'
        #agent.aeroport = request.form.get('aeroport') == 'on'
        agent.patrouille_cotiere = request.form.get('patrouille_cotiere') == 'on'
        agent.inspection_usine = request.form.get('inspection_usine') == 'on'
        agent.patrouille_aerienne = request.form.get('patrouille_aerienne') == 'on'
        agent.gardiennage = request.form.get('gardiennage') == 'on'
        agent.courrier = request.form.get('courrier') == 'on'
        agent.missions = request.form.get('missions') == 'on'
        agent.crss = request.form.get('crss') == 'on'
        agent.bvp = request.form.get('bvp') == 'on'
        agent.certification_dpsp = request.form.get('certification_dpsp') == 'on'
        agent.certification = request.form.get('certification') == 'on'
        # =================================================
        
        password = request.form.get('password')
        if password:
            agent.set_password(password)
        
        db.session.commit()
        flash(f'Agent {agent.prenom} {agent.nom} modifié avec succès!', 'success')
        return redirect(url_for('gestion_agents'))
    
    return render_template('modifier_agent.html', agent=agent)


@app.route('/admin/agents/<int:agent_id>/supprimer', methods=['POST'])
@login_required
def supprimer_agent(agent_id):
    """Supprimer un agent"""
    if current_user.role != 'admin':
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard_agent'))
    
    agent = User.query.get_or_404(agent_id)
    nom_complet = agent.nom_complet
    
    db.session.delete(agent)
    db.session.commit()
    
    flash(f'Agent {nom_complet} supprimé avec succès!', 'success')
    return redirect(url_for('gestion_agents'))


# ==================== GESTION DES PLANNINGS ====================

@app.route('/admin/plannings')
@login_required
def liste_plannings():
    """Liste de tous les plannings"""
    if current_user.role != 'admin':
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard_agent'))
    
    plannings = Planning.query.order_by(Planning.date_debut.desc()).all()
    return render_template('liste_plannings.html', plannings=plannings)


@app.route('/admin/plannings/<int:planning_id>')
@login_required
def voir_planning(planning_id):
    """Affiche les détails d'un planning avec toutes les activités"""
    planning = Planning.query.get_or_404(planning_id)
    
    # Récupérer toutes les affectations
    affectations = Affectation.query.filter_by(planning_id=planning.id).order_by(Affectation.jour, Affectation.shift).all()
    
    # Organiser par jour
    jours_semaine = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    affectations_par_jour = {}
    
    for i in range(7):
        jour_date = planning.date_debut + timedelta(days=i)
        nom_jour = jours_semaine[i]
        
        jour_affectations = [aff for aff in affectations if aff.jour == jour_date]
        
        affectations_par_jour[nom_jour] = {
            'date': jour_date,
            'affectations': jour_affectations
        }
    
    # Organiser par activités pour l'affichage
    activites_ordre = [
        ('VEILLE_CRSS', 'I - VEILLE CRSS', 'yellow'),
        ('BVP', 'II - BRIGADE DE VEILLE PORTUAIRE', 'orange'),
        ('CHAUFFEUR', 'Chauffeurs BVP', 'yellow'),
        ('REPOS_CHAUFFEUR', 'III - REPOS CHAUFFEURS', 'green'),
        ('CERTIFICATION_DIRECTION', 'IV - CERTIFICATION CAPTURE - DIRECTION', 'blue'),
        ('CERTIFICATION_AEROPORT', 'V - CERTIFICATION CAPTURE - AEROPORT', 'teal'),
        ('PATROUILLE_COTIERE', 'VI - PATROUILLE MARITIME COTIERE', 'indigo'),
        ('GARDIENNAGE', 'VII - GARDIENNAGE', 'purple'),
        ('COURRIER', 'VIII - COURRIER', 'pink'),
        ('INSPECTION_USINE', 'IX - INSPECTION USINES', 'green'),
        ('PATROUILLE_AERIENNE', 'X - PATROUILLE AERIENNE', 'blue'),
    ]
    
    return render_template('voir_planning.html',
                         planning=planning,
                         affectations_par_jour=affectations_par_jour,
                         jours_semaine=jours_semaine,
                         activites_ordre=activites_ordre)
# def voir_planning(planning_id):
#     """Voir un planning détaillé"""
#     planning = Planning.query.get_or_404(planning_id)
    
#     affectations_par_jour = {}
#     jours_semaine = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
#     for i in range(7):
#         jour = planning.date_debut + timedelta(days=i)
#         affectations_jour = Affectation.query.filter_by(
#             planning_id=planning.id,
#             jour=jour
#         ).all()
        
#         affectations_par_jour[jours_semaine[i]] = {
#             'date': jour,
#             'affectations': affectations_jour
#         }
    
#     return render_template('voir_planning.html', 
#                          planning=planning,
#                          affectations_par_jour=affectations_par_jour,
#                          jours_semaine=jours_semaine)


@app.route('/admin/plannings/generer', methods=['POST'])
@login_required
def generer_planning():
    """Générer automatiquement un nouveau planning"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    try:
        planning = scheduler_manager.generer_planning_semaine()
        # ✅ fix: recharger l'objet pour le rattacher à la session active
        planning = Planning.query.get(planning.id)
        flash(f'Planning genere pour la semaine du {planning.date_debut.strftime("%d/%m/%Y")}!', 'success')
        return redirect(url_for('voir_planning', planning_id=planning.id))
    except Exception as e:
        print(f'Erreur lors de la génération du planning: {str(e)}', 'danger')
        return redirect(url_for('liste_plannings'))


@app.route('/admin/plannings/<int:planning_id>/archiver', methods=['POST'])
@login_required
def archiver_planning(planning_id):
    """Archiver un planning"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    planning = Planning.query.get_or_404(planning_id)
    planning.statut = 'archive'
    db.session.commit()
    
    flash('Planning archivé avec succès!', 'success')
    return redirect(url_for('liste_plannings'))

@app.route('/admin/plannings/<int:planning_id>/supprimer', methods=['POST'])
@login_required
def supprimer_planning(planning_id):
    """Supprimer un planning"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    planning = Planning.query.get_or_404(planning_id)
    
    # Supprimer d'abord toutes les affectations liées
    Affectation.query.filter_by(planning_id=planning.id).delete()
    
    # Puis supprimer le planning
    db.session.delete(planning)
    db.session.commit()
    
    flash('Planning supprimé avec succès!', 'success')
    return redirect(url_for('liste_plannings'))


# ==================== EXPORT ====================

@app.route('/admin/plannings/<int:planning_id>/export/pdf')
@login_required
def export_planning_pdf(planning_id):
    """Exporter un planning en PDF - Format identique à l'Excel avec en-tête institutionnel"""
    planning = Planning.query.get_or_404(planning_id)
    
    buffer = BytesIO()
    # Marges réduites pour maximiser l'espace
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4), 
        leftMargin=10, 
        rightMargin=10, 
        topMargin=10, 
        bottomMargin=10
    )
    elements = []
    
    styles = getSampleStyleSheet()
    style_center = ParagraphStyle(
        name='center', 
        alignment=TA_CENTER, 
        fontSize=7,
        leading=9
    )
    style_title = ParagraphStyle(
        name='title',
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=4,
        fontName='Helvetica-Bold'
    )

    # ========== EN-TÊTE INSTITUTIONNEL (Compact) ==========
    logo_img = None
    sig_img = None
    
    try:
        with open("static/img/dpsp.png", "rb") as f:
            logo_bytes = f.read()
            logo_img = Image(BytesIO(logo_bytes))
            logo_img.drawHeight = 30
            logo_img.drawWidth = 30
    except Exception:
        logo_img = None

    try:
        with open("static/img/senegal.jpg", "rb") as f:
            sig_bytes = f.read()
            sig_img = Image(BytesIO(sig_bytes))
            sig_img.drawHeight = 18
            sig_img.drawWidth = 18
    except Exception:
        sig_img = None

    haut_text = (
        "<strong>République du Sénégal</strong><br/>"
        "<em>Un Peuple – Un But – Une Foi</em><br/>"
    )
    haut_para = Paragraph(haut_text, style_center)

    bas_text = (
        "★★★★★<br/>"
        "MINISTÈRE DES PÊCHES ET DE L'ÉCONOMIE MARITIME<br/>"
        "★★★★★<br/>Direction de la Protection et de la Surveillance des Pêches (DPSP)"
    )
    bas_para = Paragraph(bas_text, style_center)

    if sig_img:
        gauche_col = Table(
            [[haut_para], [sig_img], [bas_para]],
            colWidths=[doc.width/3 - 5]
        )
        gauche_col.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))
    else:
        gauche_col = Paragraph(haut_text + "★★★★★<br/>" + bas_text, style_center)

    centre_vide = Paragraph("", style_center)

    if logo_img:
        droite_col = Table([[logo_img]], colWidths=[doc.width/3 - 5])
        droite_col.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
    else:
        droite_col = Paragraph("", style_center)

    entete_table = Table(
        [[gauche_col, centre_vide, droite_col]],
        colWidths=[doc.width/3, doc.width/3, doc.width/3]
    )
    entete_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(entete_table)
    elements.append(Spacer(1, 4))

    # ========== TITRE DU PLANNING ==========
    date_debut_str = planning.date_debut.strftime('%d/%m/%Y')
    date_fin_str = planning.date_fin.strftime('%d/%m/%Y')
    mois_annee = planning.date_debut.strftime('%B %Y').upper()
    
    titre = Paragraph(
        f"<b>PLANNING HEBDOMADAIRE DE TRAVAIL DU {date_debut_str.split('/')[0]} AU {date_fin_str.split('/')[0]} {mois_annee}</b>",
        style_title
    )
    elements.append(titre)
    elements.append(Spacer(1, 4))

    # ========== CONSTRUIRE LES DONNÉES DU TABLEAU ==========
    jours_semaine = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    # En-tête des dates
    dates_header = ['Dates']
    for i in range(7):
        jour_date = planning.date_debut + timedelta(days=i)
        dates_header.append(f"{jours_semaine[i]}\n{jour_date.strftime('%d/%m/%Y')}")
    
    data = [dates_header]
    row_styles = []
    current_row = 1
    
    # Fonction helper pour ajouter une ligne de section
    def ajouter_titre_section(titre, couleur_hex):
        nonlocal current_row
        data.append([titre] + [''] * 7)
        row_styles.append({
            'row': current_row,
            'type': 'section',
            'color': colors.HexColor(f'#{couleur_hex}')
        })
        current_row += 1
    
    # Fonction helper pour ajouter une ligne de poste avec nouveaux champs
    def ajouter_ligne_poste(label, activite, sous_activite_filter=None, shift_filter=None, label_color_hex=None):
        nonlocal current_row
        
        row_data = [label]
        
        # Pour chaque jour
        for i in range(7):
            jour = planning.date_debut + timedelta(days=i)
            
            # Chercher les affectations
            query = Affectation.query.filter_by(
                planning_id=planning.id,
                jour=jour,
                activite=activite
            )
            
            if shift_filter:
                query = query.filter_by(shift=shift_filter)
            
            affs = query.all()
            
            # Filtrer par sous_activite si nécessaire
            if sous_activite_filter and affs:
                affs = [aff for aff in affs if aff.sous_activite and sous_activite_filter in aff.sous_activite]
            
            # Récupérer les noms
            noms = [aff.agent.nom_complet for aff in affs]
            row_data.append('\n'.join(noms) if noms else '')
        
        data.append(row_data)
        
        if label_color_hex:
            row_styles.append({
                'row': current_row,
                'type': 'label_color',
                'color': colors.HexColor(f'#{label_color_hex}')
            })
        
        current_row += 1
    
    # ========== SECTION I - VEILLE CRSS ==========
    ajouter_titre_section('I - VEILLE CRSS', '70AD47')
    ajouter_ligne_poste('Journée:  08h - 17h', 'VEILLE_CRSS', shift_filter='jour', label_color_hex='FFFF00')
    ajouter_ligne_poste('Nuit : 17h - 08h', 'VEILLE_CRSS', shift_filter='nuit', label_color_hex='00B050')
    
    # ========== SECTION II - BRIGADE DE VEILLE PORTUAIRE ==========
    ajouter_titre_section('II - BRIGADE DE VEILLE PORTUAIRE', 'FFC000')
    ajouter_ligne_poste("Chef d'equipe", 'BVP', sous_activite_filter='Chef')
    ajouter_ligne_poste('Inspecteur de Veille', 'BVP', sous_activite_filter='Inspecteur')
    ajouter_ligne_poste('Autres agents', 'BVP', sous_activite_filter='Autres')
    
    # Ligne vide
    data.append([''] * 8)
    current_row += 1
    
    ajouter_ligne_poste('Chauffeurs', 'CHAUFFEUR', label_color_hex='FFFF00')
    
    # ========== SECTION III - REPOS CHAUFFEURS ==========
    ajouter_titre_section('III - REPOS CHAUFFEURS', '92D050')
    ajouter_ligne_poste('Chauffeurs', 'REPOS_CHAUFFEUR')
    
    # ========== SECTION IV - CERTIFICATION DIRECTION ==========
    ajouter_titre_section('IV - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - DIRECTION', '00B0F0')
    
    # Agents de certification direction (peut y en avoir plusieurs par jour)
    row_data = ['Agents']
    for i in range(7):
        jour = planning.date_debut + timedelta(days=i)
        affs = Affectation.query.filter_by(
            planning_id=planning.id,
            jour=jour,
            activite='CERTIFICATION_DIRECTION'
        ).all()
        
        # Prendre le premier agent
        row_data.append(affs[0].agent.nom_complet if affs else '')
    data.append(row_data)
    current_row += 1
    
    # Ligne 2 des agents
    row_data = ['']
    for i in range(7):
        jour = planning.date_debut + timedelta(days=i)
        affs = Affectation.query.filter_by(
            planning_id=planning.id,
            jour=jour,
            activite='CERTIFICATION_DIRECTION'
        ).all()
        row_data.append(affs[1].agent.nom_complet if len(affs) > 1 else '')
    data.append(row_data)
    current_row += 1
    
    # Ligne 3 des agents
    row_data = ['']
    for i in range(7):
        jour = planning.date_debut + timedelta(days=i)
        affs = Affectation.query.filter_by(
            planning_id=planning.id,
            jour=jour,
            activite='CERTIFICATION_DIRECTION'
        ).all()
        row_data.append(affs[2].agent.nom_complet if len(affs) > 2 else '')
    data.append(row_data)
    current_row += 1
    
    # ========== SECTION V - CERTIFICATION AÉROPORT ==========
    ajouter_titre_section('V - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - AEROPORT', '00B050')
    ajouter_ligne_poste('Agents', 'CERTIFICATION_AEROPORT')
    
    # ========== SECTION VI - PATROUILLE MARITIME ==========
    ajouter_titre_section('VI - PATROUILLE MARITIME COTIERE', '0070C0')
    ajouter_ligne_poste('Inspecteurs', 'PATROUILLE_COTIERE')
    
    # ========== SECTION VII - GARDIENNAGE ==========
    ajouter_titre_section('VII - GARDIENNAGE', '7030A0')
    ajouter_ligne_poste('Agents', 'GARDIENNAGE')
    
    # ========== SECTION VIII - COURRIER ==========
    ajouter_titre_section('VIII - COURRIER', 'FFC000')
    ajouter_ligne_poste('Chauffeurs', 'COURRIER', label_color_hex='FFFF00')
    
    # ========== SECTION IX - INSPECTION USINES ==========
    ajouter_titre_section('IX - INSPECTION USINES', '00B050')
    
    # Inspecteurs (peut y en avoir plusieurs)
    row_data = ['Inspecteurs']
    for i in range(7):
        jour = planning.date_debut + timedelta(days=i)
        affs = Affectation.query.filter_by(
            planning_id=planning.id,
            jour=jour,
            activite='INSPECTION_USINE'
        ).filter(Affectation.sous_activite.like('%Inspecteur%')).all()
        
        noms = [aff.agent.nom_complet for aff in affs]
        row_data.append('\n'.join(noms) if noms else '')
    data.append(row_data)
    current_row += 1
    
    # Ligne vide
    data.append([''] * 8)
    current_row += 1
    
    ajouter_ligne_poste('Chauffeurs', 'INSPECTION_USINE', sous_activite_filter='Chauffeur', label_color_hex='FFFF00')
    
    # ========== SECTION X - PATROUILLE AÉRIENNE ==========
    ajouter_titre_section('X - PATROUILLE AERIENNE', '4472C4')
    ajouter_ligne_poste('Inspecteurs', 'PATROUILLE_AERIENNE')
    
    # ========== SECTION XI - MISSION/CONGÉS ==========
    ajouter_titre_section('XI - EN MISSION OU EN CONGES', 'C00000')
    
    # En-tête INSPECTEURS
    data.append(['INSPECTEURS', '', '', '', '', 'PERIODE', '', ''])
    row_styles.append({
        'row': current_row,
        'type': 'subheader',
        'color': colors.white
    })
    current_row += 1
    
    # Chercher les agents en mission
    agents_en_mission = User.query.filter_by(role='agent', est_absent=True, missions=True).all()
    
    if agents_en_mission:
        for idx, agent in enumerate(agents_en_mission[:3], 1):
            data.append([f"{idx}. {agent.nom_complet}", '', '', '', '', 'En Mission', '', ''])
            current_row += 1
    else:
        # Lignes vides
        for _ in range(3):
            data.append([''] * 8)
            current_row += 1
    
    # Ligne vide
    data.append([''] * 8)
    current_row += 1
    
    # CHAUFFEURS
    data.append(['CHAUFFEURS'] + [''] * 7)
    current_row += 1
    data.append([''] * 8)
    current_row += 1
    
    # ========== SECTION XII - EMBARQUEMENTS ==========
    ajouter_titre_section('XII -  EMBARQUEMENTS', '002060')
    
    # En-tête
    data.append(['OBSERVATEURS', '', '', '', '', 'PERIODE', '', ''])
    row_styles.append({
        'row': current_row,
        'type': 'subheader',
        'color': colors.white
    })
    current_row += 1
    
    # Chercher les observateurs embarqués
    observateurs = User.query.filter(User.date_embarquement.isnot(None)).all()
    obs_actifs = []
    
    for obs in observateurs:
        if obs.date_embarquement and obs.date_debarquement_prevue:
            if (obs.date_embarquement <= planning.date_fin and 
                obs.date_debarquement_prevue >= planning.date_debut):
                obs_actifs.append(obs)
    
    if obs_actifs:
        for idx, obs in enumerate(obs_actifs[:3], 1):
            periode = ''
            if obs.date_embarquement and obs.date_debarquement_prevue:
                periode = f"{obs.date_embarquement.strftime('%d/%m/%Y')} - {obs.date_debarquement_prevue.strftime('%d/%m/%Y')}"
            data.append([f"{idx}. {obs.nom_complet}", '', '', '', '', periode, '', ''])
            current_row += 1
    else:
        for _ in range(3):
            data.append([''] * 8)
            current_row += 1
    
    # ========== CRÉER LE TABLEAU AVEC STYLES ==========
    col_widths = [doc.width * 0.16] + [doc.width * 0.12] * 7
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Styles de base
    base_styles = [
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFFF00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6.5),
        ('FONTSIZE', (0, 1), (-1, -1), 5.5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]
    
    # Appliquer les styles spéciaux
    for style_info in row_styles:
        row_idx = style_info['row']
        
        if style_info['type'] == 'section':
            # Titre de section : fond coloré, texte noir, fusionné
            base_styles.extend([
                ('BACKGROUND', (0, row_idx), (-1, row_idx), style_info['color']),
                ('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.black),
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, row_idx), (-1, row_idx), 6.5),
                ('SPAN', (0, row_idx), (-1, row_idx)),
            ])
        
        elif style_info['type'] == 'label_color':
            # Label coloré (première colonne seulement)
            base_styles.extend([
                ('BACKGROUND', (0, row_idx), (0, row_idx), style_info['color']),
                ('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.black),
                ('FONTNAME', (0, row_idx), (0, row_idx), 'Helvetica-Bold'),
            ])
        
        elif style_info['type'] == 'subheader':
            # Sous-en-tête (INSPECTEURS, OBSERVATEURS)
            base_styles.extend([
                ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#E0E0E0')),
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, row_idx), (-1, row_idx), 6),
            ])
    
    table.setStyle(TableStyle(base_styles))
    
    elements.append(table)
    
    # ========== GÉNÉRER LE PDF ==========
    doc.build(elements)
    
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=planning_semaine_{planning.semaine}_{planning.annee}.pdf'
    return response
# @app.route('/admin/plannings/<int:planning_id>/export/excel')
# @login_required
# def export_planning_excel(planning_id):
#     """Exporter un planning en Excel - Format exact du template"""
#     planning = Planning.query.get_or_404(planning_id)
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = f"Semaine {planning.semaine}"
    
#     # ========== TITRE DU PLANNING ==========
#     date_debut_str = planning.date_debut.strftime('%d/%m/%Y')
#     date_fin_str = planning.date_fin.strftime('%d/%m/%Y')
#     mois_annee = planning.date_debut.strftime('%B %Y').upper()
    
#     ws.merge_cells('A1:H1')
#     title_cell = ws['A1']
#     title_cell.value = f"PLANNING HEBDOMADAIRE DE TRAVAIL DU {date_debut_str.split('/')[0]} AU {date_fin_str.split('/')[0]} {mois_annee}"
#     title_cell.font = Font(bold=True, size=14, color="000000")
#     title_cell.alignment = Alignment(horizontal='center', vertical='center')
#     title_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#     ws.row_dimensions[1].height = 30
    
#     # ========== EN-TÊTE DES JOURS ==========
#     jours_semaine = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
#     header_row = ['Dates']
    
#     for i in range(7):
#         jour_date = planning.date_debut + timedelta(days=i)
#         header_row.append(f"{jours_semaine[i]} {jour_date.strftime('%d/%m/%Y')}")
    
#     ws.append(header_row)
    
#     # Style de l'en-tête
#     for col_num, cell in enumerate(ws[2], 1):
#         cell.font = Font(bold=True, color="000000", size=9)
#         cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border = Border(
#             left=Side(style='thin'), right=Side(style='thin'),
#             top=Side(style='thin'), bottom=Side(style='thin')
#         )
#     ws.row_dimensions[2].height = 25
    
#     current_row = 3
#     border_thin = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )
    
#     # ========== ORGANISER LES DONNÉES ==========
#     sections_data = _organiser_affectations_par_sections(planning)
    
#     # ========== SECTION I - VEILLE CRSS ==========
#     section = sections_data['I_VEILLE_CRSS']
#     ws.append([section['titre']] + [''] * 7)
#     ws.merge_cells(f'A{current_row}:H{current_row}')
#     cell = ws[f'A{current_row}']
#     cell.font = Font(bold=True, color="000000", size=10)
#     cell.fill = PatternFill(start_color=section['couleur'], end_color=section['couleur'], fill_type="solid")
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#     cell.border = border_thin
#     ws.row_dimensions[current_row].height = 20
#     current_row += 1
    
#     # Sous-section Journée
#     journee = section['sous_sections']['journee']
    
#     # UNE SEULE LIGNE : Label + Noms des agents pour chaque jour
#     row_data = [journee['titre']]
#     for i in range(7):  # Pour chaque jour
#         jour = planning.date_debut + timedelta(days=i)
#         # Trouver l'agent affecté ce jour
#         aff = Affectation.query.filter_by(
#             planning_id=planning.id,
#             jour=jour,
#             shift='jour',
#             equipe='CRSS'
#         ).first()
        
#         if aff:
#             row_data.append(aff.agent.nom_complet)
#         else:
#             row_data.append('')
    
#     ws.append(row_data)
    
#     # Styler la ligne
#     for col in range(1, 9):
#         cell = ws.cell(row=current_row, column=col)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.border = border_thin
#         if col == 1:  # Label
#             cell.font = Font(bold=True, size=9)
#             cell.fill = PatternFill(start_color=journee['couleur'], end_color=journee['couleur'], fill_type="solid")
#         else:  # Noms agents
#             cell.font = Font(size=9)
#     current_row += 1
    
#     # Sous-section Nuit
#     nuit = section['sous_sections']['nuit']
    
#     # UNE SEULE LIGNE : Label + Noms des agents pour chaque jour
#     row_data = [nuit['titre']]
#     for i in range(7):  # Pour chaque jour
#         jour = planning.date_debut + timedelta(days=i)
#         # Trouver l'agent affecté ce jour
#         aff = Affectation.query.filter_by(
#             planning_id=planning.id,
#             jour=jour,
#             shift='nuit',
#             equipe='CRSS'
#         ).first()
        
#         if aff:
#             row_data.append(aff.agent.nom_complet)
#         else:
#             row_data.append('')
    
#     ws.append(row_data)
    
#     # Styler la ligne
#     for col in range(1, 9):
#         cell = ws.cell(row=current_row, column=col)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.border = border_thin
#         if col == 1:  # Label
#             cell.font = Font(bold=True, size=9)
#             cell.fill = PatternFill(start_color=nuit['couleur'], end_color=nuit['couleur'], fill_type="solid")
#         else:  # Noms agents
#             cell.font = Font(size=9)
#     current_row += 1
    
#     # ========== SECTION II - BRIGADE BVP ==========
#     section = sections_data['II_BRIGADE_BVP']
#     ws.append([section['titre']] + [''] * 7)
#     ws.merge_cells(f'A{current_row}:H{current_row}')
#     cell = ws[f'A{current_row}']
#     cell.font = Font(bold=True, color="000000", size=10)
#     cell.fill = PatternFill(start_color=section['couleur'], end_color=section['couleur'], fill_type="solid")
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#     cell.border = border_thin
#     ws.row_dimensions[current_row].height = 20
#     current_row += 1
    
#     # Pour chaque poste dans la brigade
#     for poste_key in ['chef', 'inspecteur', 'autres']:
#         poste_data = section['postes'][poste_key]
        
#         # UNE SEULE LIGNE : Label du poste + Noms des agents pour chaque jour
#         row_data = [poste_data['titre']]
#         for i in range(7):  # Pour chaque jour
#             jour = planning.date_debut + timedelta(days=i)
#             # Trouver l'agent affecté ce jour pour ce poste
#             aff = Affectation.query.filter_by(
#                 planning_id=planning.id,
#                 jour=jour,
#                 equipe='BVP',
#                 poste=poste_key
#             ).first()
            
#             if aff:
#                 row_data.append(aff.agent.nom_complet)
#             else:
#                 row_data.append('')
        
#         ws.append(row_data)
        
#         # Styler la ligne
#         for col in range(1, 9):
#             cell = ws.cell(row=current_row, column=col)
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#             cell.border = border_thin
#             cell.font = Font(size=9)
#             if col == 1:  # Label en gras
#                 cell.font = Font(bold=True, size=9)
#         current_row += 1
    
#     # Ligne vide
#     ws.append([''] * 8)
#     current_row += 1
    
#     # Chauffeurs (avec fond jaune) - UNE SEULE LIGNE
#     chauffeurs = section['postes']['chauffeurs']
#     row_data = [chauffeurs['titre']]
    
#     for i in range(7):  # Pour chaque jour
#         jour = planning.date_debut + timedelta(days=i)
#         # Trouver le chauffeur affecté ce jour
#         aff = Affectation.query.filter_by(
#             planning_id=planning.id,
#             jour=jour,
#             equipe='BVP',
#             poste='chauffeur'
#         ).first()
        
#         if aff:
#             row_data.append(aff.agent.nom_complet)
#         else:
#             row_data.append('')
    
#     ws.append(row_data)
    
#     for col in range(1, 9):
#         cell = ws.cell(row=current_row, column=col)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.border = border_thin
#         if col == 1:  # Label
#             cell.font = Font(bold=True, size=9)
#             cell.fill = PatternFill(start_color=chauffeurs['couleur'], end_color=chauffeurs['couleur'], fill_type="solid")
#         else:  # Noms
#             cell.font = Font(size=9)
#     current_row += 1
    
#     # ========== SECTIONS III À X - Même pattern ==========
#     # Pour gagner du temps, je crée une fonction helper
#     def ajouter_section_simple(section_key, has_subsection=False, subsection_keys=None):
#         nonlocal current_row
#         section = sections_data[section_key]
        
#         # Titre de section
#         ws.append([section['titre']] + [''] * 7)
#         ws.merge_cells(f'A{current_row}:H{current_row}')
#         cell = ws[f'A{current_row}']
#         cell.font = Font(bold=True, color="000000", size=10)
#         cell.fill = PatternFill(start_color=section['couleur'], end_color=section['couleur'], fill_type="solid")
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.border = border_thin
#         ws.row_dimensions[current_row].height = 20
#         current_row += 1
        
#         if has_subsection and subsection_keys:
#             for sub_key in subsection_keys:
#                 sub_data = section[sub_key] if isinstance(section.get(sub_key), dict) else None
#                 if sub_data and 'agents' in sub_data:
#                     # Titre sous-section
#                     ws.append([sub_data.get('titre', sub_key)] + [''] * 7)
#                     for col in range(1, 9):
#                         cell = ws.cell(row=current_row, column=col)
#                         cell.font = Font(size=9)
#                         if 'couleur' in sub_data:
#                             cell.fill = PatternFill(start_color=sub_data['couleur'], 
#                                                     end_color=sub_data['couleur'], fill_type="solid")
#                         cell.alignment = Alignment(horizontal='center', vertical='center')
#                         cell.border = border_thin
#                     current_row += 1
                    
#                     # Agents
#                     for agent_id, agent_data in sub_data['agents'].items():
#                         row_data = [agent_data['nom']]
#                         for i in range(7):
#                             if i in agent_data['jours']:
#                                 row_data.append(agent_data['jours'][i].agent.nom_complet)
#                             else:
#                                 row_data.append('')
#                         ws.append(row_data)
                        
#                         for col in range(1, 9):
#                             cell = ws.cell(row=current_row, column=col)
#                             cell.alignment = Alignment(horizontal='center', vertical='center')
#                             cell.border = border_thin
#                             cell.font = Font(size=9)
#                         current_row += 1
                    
#                     if not sub_data['agents']:
#                         ws.append([''] * 8)
#                         current_row += 1
#         else:
#             # Section simple avec juste agents
#             if 'agents' in section:
#                 for agent_id, agent_data in section['agents'].items():
#                     row_data = [agent_data['nom']]
#                     for i in range(7):
#                         if i in agent_data['jours']:
#                             row_data.append(agent_data['jours'][i].agent.nom_complet)
#                         else:
#                             row_data.append('')
#                     ws.append(row_data)
                    
#                     for col in range(1, 9):
#                         cell = ws.cell(row=current_row, column=col)
#                         cell.alignment = Alignment(horizontal='center', vertical='center')
#                         cell.border = border_thin
#                         cell.font = Font(size=9)
#                     current_row += 1
                
#                 if not section['agents']:
#                     ws.append([''] * 8)
#                     current_row += 1
    
#     # Ajouter toutes les sections restantes
#     ajouter_section_simple('III_REPOS_CHAUFFEURS', has_subsection=True, subsection_keys=['chauffeurs'])
#     ajouter_section_simple('IV_CERTIFICATION_DIR')
#     ajouter_section_simple('V_CERTIFICATION_AERO')
#     ajouter_section_simple('VI_PATROUILLE_MARITIME', has_subsection=True, subsection_keys=['inspecteurs'])
#     ajouter_section_simple('VII_GARDIENNAGE')
#     ajouter_section_simple('VIII_COURRIER', has_subsection=True, subsection_keys=['chauffeurs'])
    
#     # Section IX avec postes multiples
#     section = sections_data['IX_INSPECTION_USINES']
#     ws.append([section['titre']] + [''] * 7)
#     ws.merge_cells(f'A{current_row}:H{current_row}')
#     cell = ws[f'A{current_row}']
#     cell.font = Font(bold=True, color="000000", size=10)
#     cell.fill = PatternFill(start_color=section['couleur'], end_color=section['couleur'], fill_type="solid")
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#     cell.border = border_thin
#     ws.row_dimensions[current_row].height = 20
#     current_row += 1
    
#     for poste_key in ['inspecteurs', 'chauffeurs']:
#         poste_data = section['postes'][poste_key]
#         ws.append([poste_data['titre']] + [''] * 7)
#         for col in range(1, 9):
#             cell = ws.cell(row=current_row, column=col)
#             cell.font = Font(size=9)
#             if 'couleur' in poste_data:
#                 cell.fill = PatternFill(start_color=poste_data['couleur'], 
#                                         end_color=poste_data['couleur'], fill_type="solid")
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#             cell.border = border_thin
#         current_row += 1
        
#         for agent_id, agent_data in poste_data['agents'].items():
#             row_data = [agent_data['nom']]
#             for i in range(7):
#                 if i in agent_data['jours']:
#                     row_data.append(agent_data['jours'][i].agent.nom_complet)
#                 else:
#                     row_data.append('')
#             ws.append(row_data)
            
#             for col in range(1, 9):
#                 cell = ws.cell(row=current_row, column=col)
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 cell.border = border_thin
#                 cell.font = Font(size=9)
#             current_row += 1
        
#         if not poste_data['agents']:
#             ws.append([''] * 8)
#             current_row += 1
    
#     ajouter_section_simple('X_PATROUILLE_AERIENNE', has_subsection=True, subsection_keys=['inspecteurs'])
    
#     # ========== SECTION XI - MISSION/CONGES (Format spécial) ==========
#     section = sections_data['XI_MISSION_CONGES']
#     ws.append([section['titre']] + [''] * 7)
#     ws.merge_cells(f'A{current_row}:H{current_row}')
#     cell = ws[f'A{current_row}']
#     cell.font = Font(bold=True, color="000000", size=10)
#     cell.fill = PatternFill(start_color=section['couleur'], end_color=section['couleur'], fill_type="solid")
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#     cell.border = border_thin
#     ws.row_dimensions[current_row].height = 20
#     current_row += 1
    
#     # En-tête spécial pour cette section
#     ws.append(['PRENOMS ET NOM', '', '', '', '', 'PERIODE', '', ''])
#     ws.merge_cells(f'A{current_row}:E{current_row}')
#     ws.merge_cells(f'F{current_row}:H{current_row}')
#     for col in range(1, 9):
#         cell = ws.cell(row=current_row, column=col)
#         cell.font = Font(bold=True, size=9)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.border = border_thin
#     current_row += 1
    
#     # Personnes en mission/congés
#     for idx, personne in enumerate(section['personnes'], 1):
#         ws.append([f"{idx}. {personne['nom']}", '', '', '', '', personne['periode'], '', ''])
#         ws.merge_cells(f'A{current_row}:E{current_row}')
#         ws.merge_cells(f'F{current_row}:H{current_row}')
#         for col in range(1, 9):
#             cell = ws.cell(row=current_row, column=col)
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#             cell.border = border_thin
#             cell.font = Font(size=9)
#         current_row += 1
    
#     if not section['personnes']:
#         for _ in range(3):  # 3 lignes vides
#             ws.append([''] * 8)
#             current_row += 1
    
#     # Ajouter sections pour chauffeurs aussi
#     ws.append(['CHAUFFEURS', '', '', '', '', '', '', ''])
#     current_row += 1
#     ws.append([''] * 8)
#     current_row += 1
    
#     # ========== SECTION XII - EMBARQUEMENTS (Liste) ==========
#     section = sections_data['XII_EMBARQUEMENTS']
#     ws.append([section['titre']] + [''] * 7)
#     ws.merge_cells(f'A{current_row}:H{current_row}')
#     cell = ws[f'A{current_row}']
#     cell.font = Font(bold=True, color="FFFFFF", size=10)
#     cell.fill = PatternFill(start_color=section['couleur'], end_color=section['couleur'], fill_type="solid")
#     cell.alignment = Alignment(horizontal='center', vertical='center')
#     cell.border = border_thin
#     ws.row_dimensions[current_row].height = 20
#     current_row += 1
    
#     # En-tête pour les observateurs
#     ws.append(['PRENOMS ET NOM', '', '', '', '', 'PERIODE', '', ''])
#     ws.merge_cells(f'A{current_row}:E{current_row}')
#     ws.merge_cells(f'F{current_row}:H{current_row}')
#     for col in range(1, 9):
#         cell = ws.cell(row=current_row, column=col)
#         cell.font = Font(bold=True, size=9)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.border = border_thin
#     current_row += 1
    
#     # Liste des observateurs embarqués
#     for idx, obs in enumerate(section['observateurs'], 1):
#         periode = ''
#         if obs.get('date_emb') and obs.get('date_deb'):
#             periode = f"{obs['date_emb'].strftime('%d/%m/%Y')} - {obs['date_deb'].strftime('%d/%m/%Y')}"
        
#         ws.append([f"{idx}. {obs['nom']}", '', '', '', '', periode, '', ''])
#         ws.merge_cells(f'A{current_row}:E{current_row}')
#         ws.merge_cells(f'F{current_row}:H{current_row}')
#         for col in range(1, 9):
#             cell = ws.cell(row=current_row, column=col)
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#             cell.border = border_thin
#             cell.font = Font(size=9, color="0000FF")  # Bleu comme dans le template
#         current_row += 1
    
#     if not section['observateurs']:
#         for _ in range(5):  # 5 lignes vides
#             ws.append([''] * 8)
#             current_row += 1
    
#     # ========== AJUSTER LES LARGEURS ==========
#     ws.column_dimensions['A'].width = 25
#     for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
#         ws.column_dimensions[col].width = 18
    
#     # ========== SAUVEGARDER ==========
#     buffer = BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)
    
#     response = make_response(buffer.getvalue())
#     response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     response.headers['Content-Disposition'] = f'attachment; filename=planning_semaine_{planning.semaine}_{planning.annee}.xlsx'
    
#     return response

@app.route('/admin/plannings/<int:planning_id>/export/excel')
@login_required
def export_planning_excel(planning_id):
    """Exporte le planning en Excel selon le format standard"""
    planning = Planning.query.get_or_404(planning_id)
    
    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"PLANNING S_{planning.semaine}"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    section_font = Font(bold=True, size=10)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Récupérer les affectations
    affectations = Affectation.query.filter_by(planning_id=planning.id).order_by(Affectation.jour).all()
    
    # Organiser par jour et activité
    affectations_par_jour = {}
    for i in range(7):
        jour_date = planning.date_debut + timedelta(days=i)
        affectations_par_jour[jour_date] = [aff for aff in affectations if aff.jour == jour_date]
    
    # ========== TITRE ==========
    ws.merge_cells('B2:I2')
    titre_cell = ws['B2']
    titre_cell.value = f"PLANNING HEBDOMADAIRE DE TRAVAIL DU {planning.date_debut.strftime('%d AU %d %B %Y').upper()}"
    titre_cell.font = Font(bold=True, size=14)
    titre_cell.alignment = center_align
    titre_cell.fill = header_fill
    ws.row_dimensions[2].height = 25
    
    # ========== EN-TÊTE DES JOURS ==========
    jours_semaine = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    ws.append(['', 'Dates'] + [f"{jour} {(planning.date_debut + timedelta(days=i)).strftime('%d/%m/%Y')}" 
                                for i, jour in enumerate(jours_semaine)])
    
    for col in range(2, 10):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = center_align
        cell.border = border_thin
    
    current_row = 4
    
    # ========== FONCTION HELPER ==========
    def ajouter_section(titre, couleur, filtrer_func):
        nonlocal current_row
        ws.append([f'', titre] + [''] * 7)
        ws.merge_cells(f'B{current_row}:I{current_row}')
        cell = ws[f'B{current_row}']
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
        cell.alignment = center_align
        cell.border = border_thin
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # Récupérer agents par jour
        agents_par_jour = {}
        for jour_date in affectations_par_jour:
            jour_idx = (jour_date - planning.date_debut).days
            agents = filtrer_func(affectations_par_jour[jour_date])
            agents_par_jour[jour_idx] = agents
        
        return agents_par_jour
    
    def ajouter_ligne_agents(label, agents_dict, couleur_cell=None):
        nonlocal current_row
        ligne = ['', label]
        for i in range(7):
            agents = agents_dict.get(i, [])
            ligne.append(', '.join([a.agent.nom_complet for a in agents]) if agents else '')
        ws.append(ligne)
        
        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_thin
            cell.alignment = center_align
            if col == 2:
                cell.font = Font(bold=True, size=9)
                if couleur_cell:
                    cell.fill = PatternFill(start_color=couleur_cell, end_color=couleur_cell, fill_type="solid")
        
        ws.row_dimensions[current_row].height = 30
        current_row += 1
    
    # ========== I - VEILLE CRSS ==========
    agents_crss_jour = ajouter_section(
        'I - VEILLE CRSS',
        '70AD47',
        lambda affs: [a for a in affs if a.activite == 'VEILLE_CRSS' and a.shift == 'jour']
    )
    ajouter_ligne_agents('Journée: 08h - 17h', agents_crss_jour, 'FFFF00')
    
    agents_crss_nuit = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_crss_nuit[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                       if a.activite == 'VEILLE_CRSS' and a.shift == 'nuit']
    ajouter_ligne_agents('Nuit : 17h - 08h', agents_crss_nuit, '00B050')
    
    # ========== II - BRIGADE BVP ==========
    ajouter_section('II - BRIGADE DE VEILLE PORTUAIRE', 'FFC000', lambda affs: [])
    
    # Chef d'équipe
    agents_bvp_chef = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_bvp_chef[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                      if a.activite == 'BVP' and "Chef" in str(a.sous_activite)]
    ajouter_ligne_agents("Chef d'equipe", agents_bvp_chef)
    
    # Inspecteur
    agents_bvp_insp = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_bvp_insp[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                      if a.activite == 'BVP' and "Inspecteur" in str(a.sous_activite)]
    ajouter_ligne_agents('Inspecteur de Veille', agents_bvp_insp)
    
    # Autres agents
    agents_bvp_autres = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_bvp_autres[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                        if a.activite == 'BVP' and "Autres" in str(a.sous_activite)]
    ajouter_ligne_agents('Autres agents', agents_bvp_autres)
    
    # Ligne vide
    ws.append([''] * 9)
    current_row += 1
    
    # Chauffeurs (ligne spéciale)
    agents_chauffeurs = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_chauffeurs[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                        if a.activite == 'CHAUFFEUR']
    ajouter_ligne_agents('Chauffeurs', agents_chauffeurs, 'FFFF00')
    
    # ========== III - REPOS CHAUFFEURS ==========
    agents_repos = ajouter_section(
        'III - REPOS CHAUFFEURS',
        '92D050',
        lambda affs: [a for a in affs if a.activite == 'REPOS_CHAUFFEUR']
    )
    ajouter_ligne_agents('Chauffeurs', agents_repos)
    
    # ========== IV - CERTIFICATION DIRECTION ==========
    agents_cert_dir = ajouter_section(
        'IV - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - DIRECTION',
        '00B0F0',
        lambda affs: [a for a in affs if a.activite == 'CERTIFICATION_DIRECTION']
    )
    ajouter_ligne_agents('Agents', agents_cert_dir)
    
    # ========== V - CERTIFICATION AEROPORT ==========
    agents_cert_aero = ajouter_section(
        'V - CERTIFICATION DE CAPTURE DES PRODUITS DE LA PECHE - AEROPORT',
        '00B050',
        lambda affs: [a for a in affs if a.activite == 'CERTIFICATION_AEROPORT']
    )
    ajouter_ligne_agents('Agents', agents_cert_aero)
    
    # ========== VI - PATROUILLE MARITIME ==========
    agents_patrol = ajouter_section(
        'VI - PATROUILLE MARITIME COTIERE',
        '0070C0',
        lambda affs: [a for a in affs if a.activite == 'PATROUILLE_COTIERE']
    )
    ajouter_ligne_agents('Inspecteurs', agents_patrol)
    
    # ========== VII - GARDIENNAGE ==========
    agents_gardien = ajouter_section(
        'VII - GARDIENNAGE',
        '7030A0',
        lambda affs: [a for a in affs if a.activite == 'GARDIENNAGE']
    )
    ajouter_ligne_agents('Agents', agents_gardien)
    
    # ========== VIII - COURRIER ==========
    agents_courrier = ajouter_section(
        'VIII - COURRIER',
        'FFC000',
        lambda affs: [a for a in affs if a.activite == 'COURRIER']
    )
    ajouter_ligne_agents('Chauffeurs', agents_courrier, 'FFFF00')
    
    # ========== IX - INSPECTION USINES ==========
    ajouter_section('IX - INSPECTION USINES', '00B050', lambda affs: [])
    
    agents_insp_usine = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_insp_usine[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                        if a.activite == 'INSPECTION_USINE' and 'Inspecteur' in str(a.sous_activite)]
    ajouter_ligne_agents('Inspecteurs', agents_insp_usine)
    
    agents_chauff_usine = {}
    for jour_date in affectations_par_jour:
        jour_idx = (jour_date - planning.date_debut).days
        agents_chauff_usine[jour_idx] = [a for a in affectations_par_jour[jour_date] 
                                          if a.activite == 'INSPECTION_USINE' and 'Chauffeur' in str(a.sous_activite)]
    ajouter_ligne_agents('Chauffeurs', agents_chauff_usine, 'FFFF00')
    
    # ========== X - PATROUILLE AERIENNE ==========
    agents_aero = ajouter_section(
        'X - PATROUILLE AERIENNE',
        '4472C4',
        lambda affs: [a for a in affs if a.activite == 'PATROUILLE_AERIENNE']
    )
    ajouter_ligne_agents('Inspecteurs', agents_aero)
    
    # Lignes vides
    ws.append([''] * 9)
    current_row += 1
    
    # ========== XI - MISSIONS/CONGES ==========
    ws.append(['', 'XI - EN MISSION OU EN CONGES'] + [''] * 7)
    ws.merge_cells(f'B{current_row}:I{current_row}')
    cell = ws[f'B{current_row}']
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = center_align
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    # Agents absents
    agents_absents = User.query.filter_by(role='agent', est_absent=True, missions=True).all()
    if agents_absents:
        ws.append(['', 'PRENOMS ET NOM', '', '', '', 'PERIODE', '', '', ''])
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws.merge_cells(f'F{current_row}:I{current_row}')
        current_row += 1
        
        for idx, agent in enumerate(agents_absents, 1):
            periode = "En Mission"
            ws.append(['', f'{idx}. {agent.nom_complet}', '', '', '', periode, '', '', ''])
            ws.merge_cells(f'B{current_row}:E{current_row}')
            ws.merge_cells(f'F{current_row}:I{current_row}')
            current_row += 1
    
    # ========== XII - EMBARQUEMENTS ==========
    ws.append([''] * 9)
    current_row += 1
    ws.append(['', 'XII - EMBARQUEMENTS'] + [''] * 7)
    ws.merge_cells(f'B{current_row}:I{current_row}')
    cell = ws[f'B{current_row}']
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    cell.alignment = center_align
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    # Observateurs embarqués (exemple)
    ws.append(['', 'PRENOMS ET NOM', '', '', '', 'PERIODE', '', '', ''])
    ws.merge_cells(f'B{current_row}:E{current_row}')
    ws.merge_cells(f'F{current_row}:I{current_row}')
    current_row += 1
    
    # ========== AJUSTER LARGEURS ==========
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 20
    
    # ========== SAUVEGARDER ==========
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=planning_S{planning.semaine}_{planning.annee}.xlsx'
    
    return response

# ==================== TÂCHES PLANIFIÉES ====================

def job_generation_automatique():
    """Job qui génère automatiquement le planning chaque dimanche"""
    print(f"[{datetime.now()}] Démarrage de la génération automatique du planning...")
    try:
        planning = scheduler_manager.generer_planning_semaine()
        print(f"[{datetime.now()}] Planning généré avec succès: Semaine {planning.semaine}/{planning.annee}")
    except Exception as e:
        print(f"[{datetime.now()}] Erreur lors de la génération automatique: {str(e)}")


# ==================== INITIALISATION ====================

@app.cli.command()
def init_db():
    """Initialise la base de données"""
    db.create_all()
    print("Base de données initialisée!")


@app.cli.command()
def create_admin():
    """Crée un compte administrateur"""
    email = input("Email: ")
    password = input("Mot de passe: ")
    nom = input("Nom: ")
    prenom = input("Prénom: ")
    
    if User.query.filter_by(email=email).first():
        print("Cet email existe déjà!")
        return
    
    admin = User(
        email=email,
        nom=nom,
        prenom=prenom,
        role='admin'
    )
    admin.set_password(password)
    
    db.session.add(admin)
    db.session.commit()
    
    print(f"Administrateur {prenom} {nom} créé avec succès!")


@app.cli.command()
def create_sample_agents():
    """Crée des agents de test"""
    agents_data = [
        {"nom": "Diop", "prenom": "Mamadou", "email": "mamadou.diop@example.com"},
        {"nom": "Ndiaye", "prenom": "Fatou", "email": "fatou.ndiaye@example.com"},
        {"nom": "Sow", "prenom": "Ibrahima", "email": "ibrahima.sow@example.com"},
        {"nom": "Fall", "prenom": "Aissatou", "email": "aissatou.fall@example.com"},
        {"nom": "Ba", "prenom": "Moussa", "email": "moussa.ba@example.com"},
        {"nom": "Gueye", "prenom": "Aminata", "email": "aminata.gueye@example.com"},
        {"nom": "Sarr", "prenom": "Omar", "email": "omar.sarr@example.com"},
        {"nom": "Sy", "prenom": "Khady", "email": "khady.sy@example.com"},
    ]
    
    for data in agents_data:
        if not User.query.filter_by(email=data["email"]).first():
            agent = User(
                nom=data["nom"],
                prenom=data["prenom"],
                email=data["email"],
                role='agent',
                disponibilite=True
            )
            agent.set_password("password123")
            db.session.add(agent)
    
    db.session.commit()
    print(f"{len(agents_data)} agents créés avec succès!")


if __name__ == '__main__':
    scheduler.add_job(
        func=job_generation_automatique,
        trigger='cron',
        day_of_week='sun',
        hour=20,
        minute=00,
        id='generation_planning_hebdomadaire'
    )
    
    scheduler.start()
    
    print("Scheduler démarré - Génération automatique chaque dimanche à 20h00")
    
    app.run(debug=True, use_reloader=False)